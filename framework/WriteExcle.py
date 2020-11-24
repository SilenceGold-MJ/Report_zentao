# -*- coding: utf-8 -*-
# !/usr/bin/python3
import openpyxl
import time
from openpyxl.styles import Font
from openpyxl import load_workbook

from openpyxl.styles import PatternFill  # 导入填充模块
from framework.logger import Logger
logger = Logger(logger="ExportExcle").getlog()

def SummaryExcle(addr,listdata):
    # 设置文件 mingc

    # 打开文件
    wb = load_workbook(addr)  # load_workbook(addr)
    # 创建一张新表
    now = time.strftime("%Y%m%d%H%M%S", time.localtime(time.time()))  # 获取当前时间
    ws = wb.create_sheet(now, 0)#插入最靠前
    #ws = wb.active
    #ws.title = '汇总页'

    sheet1 = wb[now]  # 打开第一个 sheet 工作表

    for i in listdata:
        list_row = list(i.values())
        ws.append(list_row)



    wb.save(addr)